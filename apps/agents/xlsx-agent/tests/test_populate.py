"""Tests for the synchronous /api/spreadsheets/populate endpoint."""

from __future__ import annotations

import io
import json
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.main import app


@pytest.fixture
def client() -> TestClient:
    return TestClient(app)


@pytest.fixture
def tiny_template_bytes() -> bytes:
    """A small workbook with one sheet and one input cell."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Base rate"
    ws["B1"] = None
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _minimal_cellmap() -> dict:
    return {
        "templateId": "tiny",
        "schemaVersion": "suzie-lbo-v1",
        "profile": "uk_mid_market_lbo",
        "modules": {
            "transaction_inputs": {
                "moduleVersion": "v1",
                "scalars": {"base_rate": "Sheet1!B1"},
                "slots": {},
            },
        },
        "other": [],
    }


def test_populate_requires_template_file(client: TestClient) -> None:
    """Missing template upload returns 422."""
    resp = client.post(
        "/api/spreadsheets/populate",
        data={
            "cellmap": json.dumps(_minimal_cellmap()),
            "assumptions": json.dumps({"transaction_inputs.base_rate": {"base": 0.0375}}),
        },
    )
    assert resp.status_code == 422


def test_populate_returns_xlsx(client: TestClient, tiny_template_bytes: bytes) -> None:
    """Valid request returns 200 with the populated workbook in the body."""
    cellmap = _minimal_cellmap()
    assumptions = {"transaction_inputs.base_rate": {"base": 0.0375}}
    resp = client.post(
        "/api/spreadsheets/populate",
        files={"template": ("tiny.xlsx", tiny_template_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={
            "cellmap": json.dumps(cellmap),
            "assumptions": json.dumps(assumptions),
        },
    )
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    wb = load_workbook(io.BytesIO(resp.content))
    assert wb["Sheet1"]["B1"].value == 0.0375


def test_populate_invalid_cellmap_json_returns_400(client: TestClient, tiny_template_bytes: bytes) -> None:
    """Malformed cellmap JSON returns 400."""
    resp = client.post(
        "/api/spreadsheets/populate",
        files={"template": ("tiny.xlsx", tiny_template_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={
            "cellmap": "{not json",
            "assumptions": json.dumps({}),
        },
    )
    assert resp.status_code == 400
    assert "cellmap" in resp.json().get("detail", "").lower()


def test_populate_invalid_address_returns_400(client: TestClient, tiny_template_bytes: bytes) -> None:
    """Cellmap with an address pointing at a non-existent sheet returns 400."""
    cellmap = _minimal_cellmap()
    cellmap["modules"]["transaction_inputs"]["scalars"]["base_rate"] = "Nonexistent!A1"
    resp = client.post(
        "/api/spreadsheets/populate",
        files={"template": ("tiny.xlsx", tiny_template_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={
            "cellmap": json.dumps(cellmap),
            "assumptions": json.dumps({"transaction_inputs.base_rate": {"base": 0.0375}}),
        },
    )
    assert resp.status_code == 400
    assert "nonexistent" in resp.json().get("detail", "").lower()


def test_populate_blixt_template_end_to_end(client: TestClient) -> None:
    """Populate the real shipped Blixt template using its actual cellmap."""
    fixtures_dir = Path(__file__).parent / "fixtures"
    template_bytes = (fixtures_dir / "blixt-lbo-v1.xlsx").read_bytes()
    cellmap = json.loads((fixtures_dir / "blixt-lbo-v1.cellmap.json").read_text())

    assumptions = {
        "core_meta.project_name": {"base": "Project Test"},
        "core_meta.currency": {"base": "GBP"},
        "transaction_inputs.base_rate": {"base": 0.0425},
        "transaction_inputs.tax_rate": {"base": 0.25},
        "entry_exit.entry_valuation_basis": {"base": "EBITDA"},
        "cash_flow_drivers.wc_pct_revenue_change": {"base": 0.10},
        "cash_flow_drivers.capex_maintenance_pct_revenue": {"base": 0.03},
    }

    resp = client.post(
        "/api/spreadsheets/populate",
        files={
            "template": (
                "blixt-lbo-v1.xlsx",
                template_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={
            "cellmap": json.dumps(cellmap),
            "assumptions": json.dumps(assumptions),
        },
    )
    assert resp.status_code == 200, resp.text

    wb = load_workbook(io.BytesIO(resp.content))
    # core_meta.project_name -> LBO!L13 per the cellmap
    assert wb["LBO"]["L13"].value == "Project Test"
    # transaction_inputs.base_rate -> LBO!F13
    assert wb["LBO"]["F13"].value == 0.0425
    # transaction_inputs.tax_rate -> LBO!L15
    assert wb["LBO"]["L15"].value == 0.25


def test_populate_unknown_assumption_is_ignored(client: TestClient) -> None:
    """Assumption keys not in the cellmap are silently ignored, not errors."""
    fixtures_dir = Path(__file__).parent / "fixtures"
    template_bytes = (fixtures_dir / "blixt-lbo-v1.xlsx").read_bytes()
    cellmap = json.loads((fixtures_dir / "blixt-lbo-v1.cellmap.json").read_text())

    assumptions = {
        "core_meta.project_name": {"base": "Test Project"},
        "imaginary.fake_field": {"base": 99},
    }

    resp = client.post(
        "/api/spreadsheets/populate",
        files={
            "template": ("blixt-lbo-v1.xlsx", template_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        },
        data={
            "cellmap": json.dumps(cellmap),
            "assumptions": json.dumps(assumptions),
        },
    )
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    assert wb["LBO"]["L13"].value == "Test Project"


def test_populate_rejects_non_scalar_value(client: TestClient, tiny_template_bytes: bytes) -> None:
    """A list/dict assumption value yields 400, not 500."""
    cellmap = {
        "templateId": "tiny",
        "schemaVersion": "suzie-lbo-v1",
        "profile": "uk_mid_market_lbo",
        "modules": {
            "transaction_inputs": {
                "moduleVersion": "v1",
                "scalars": {"base_rate": "Sheet1!B1"},
                "slots": {},
            },
        },
        "other": [],
    }
    assumptions = {"transaction_inputs.base_rate": {"base": [1, 2, 3]}}
    resp = client.post(
        "/api/spreadsheets/populate",
        files={"template": ("tiny.xlsx", tiny_template_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={
            "cellmap": json.dumps(cellmap),
            "assumptions": json.dumps(assumptions),
        },
    )
    assert resp.status_code == 400
    detail = resp.json().get("detail", "")
    assert "base_rate" in detail.lower() or "could not write" in detail.lower()


def test_populate_rejects_malformed_cellmap_module(client: TestClient, tiny_template_bytes: bytes) -> None:
    """A cellmap with a non-dict module value yields 400, not 500."""
    cellmap = {
        "templateId": "tiny",
        "schemaVersion": "suzie-lbo-v1",
        "profile": "uk_mid_market_lbo",
        "modules": {
            "transaction_inputs": None,
        },
        "other": [],
    }
    resp = client.post(
        "/api/spreadsheets/populate",
        files={"template": ("tiny.xlsx", tiny_template_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={
            "cellmap": json.dumps(cellmap),
            "assumptions": json.dumps({}),
        },
    )
    assert resp.status_code == 400
    assert "malformed cellmap" in resp.json().get("detail", "").lower()


def test_populate_rejects_other_missing_name(client: TestClient, tiny_template_bytes: bytes) -> None:
    """A cellmap whose `other` entry is missing required keys yields 400."""
    cellmap = {
        "templateId": "tiny",
        "schemaVersion": "suzie-lbo-v1",
        "profile": "uk_mid_market_lbo",
        "modules": {},
        "other": [{"address": "Sheet1!B1"}],  # missing "name"
    }
    resp = client.post(
        "/api/spreadsheets/populate",
        files={"template": ("tiny.xlsx", tiny_template_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={
            "cellmap": json.dumps(cellmap),
            "assumptions": json.dumps({}),
        },
    )
    assert resp.status_code == 400
    assert "malformed cellmap" in resp.json().get("detail", "").lower()
