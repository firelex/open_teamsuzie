"""Synchronous cellmap-driven workbook population.

Pure function: takes template bytes + cellmap + assumptions, returns populated .xlsx bytes.
No knowledge of the canonical schema beyond cellmap structure.
No LLM calls. No network.
"""

from __future__ import annotations

import io
import re
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook import Workbook


class PopulateError(ValueError):
    """Raised when populate_workbook cannot complete due to bad input."""


# Match "SheetName!Address" or "'Quoted Sheet'!Address".
_ADDRESS_RE = re.compile(r"^(?:'([^']+)'|([^!']+))!([A-Z]+\d+)$")


def _parse_address(address: str) -> tuple[str, str]:
    """Return (sheet_name, cell_ref) or raise PopulateError."""
    m = _ADDRESS_RE.match(address)
    if not m:
        raise PopulateError(f"invalid cell address: {address!r}")
    sheet = m.group(1) or m.group(2)
    ref = m.group(3)
    return sheet, ref


def _resolve_address(wb: Workbook, address: str) -> Any:
    """Locate the openpyxl Cell at address, raising PopulateError if sheet missing."""
    sheet_name, ref = _parse_address(address)
    if sheet_name not in wb.sheetnames:
        raise PopulateError(f"unknown sheet in address {address!r}: {sheet_name!r}")
    return wb[sheet_name][ref]


def _flatten_cellmap(cellmap: dict) -> dict[str, str]:
    """Flatten the cellmap to {canonical_path -> address} for fast lookup.

    Canonical paths:
      "<module>.<scalar>"                 e.g. "transaction_inputs.base_rate"
      "<module>.<slot>.<field>"           e.g. "debt_stack.rcf.quantum"
      "other.<name>"                      e.g. "other.sweet_equity_pct"
    """
    out: dict[str, str] = {}
    for module_id, module in cellmap.get("modules", {}).items():
        for scalar_name, addr in module.get("scalars", {}).items():
            out[f"{module_id}.{scalar_name}"] = addr
        for slot_name, slot in module.get("slots", {}).items():
            for field_name, addr in slot.get("fields", {}).items():
                out[f"{module_id}.{slot_name}.{field_name}"] = addr
    for cell in cellmap.get("other", []):
        out[f"other.{cell['name']}"] = cell["address"]
    return out


def populate_workbook(
    template_bytes: bytes,
    cellmap: dict,
    assumptions: dict[str, dict],
) -> bytes:
    """Write values from `assumptions` into `template_bytes` via `cellmap`.

    Args:
        template_bytes: raw .xlsx bytes.
        cellmap: dict matching the Cellmap shape (templateId, modules, other).
        assumptions: dict keyed by canonical path; each value is
            {"base": value, "low": value?, "high": value?}.
            Unknown keys (not present in cellmap) are ignored.
            v1: only the "base" key is written to its mapped cell.
            Low/high writes for scenario-applicable inputs land via the cellmap's
            scenarios module slots (e.g. scenarios.low.case_label).

    Returns:
        Populated .xlsx as bytes.

    Raises:
        PopulateError: on bad addresses, unknown sheets, or invalid cellmap shape.
    """
    if not isinstance(cellmap, dict) or "modules" not in cellmap:
        raise PopulateError("cellmap is missing required 'modules' key")
    if not isinstance(assumptions, dict):
        raise PopulateError("assumptions must be a dict")

    try:
        wb = load_workbook(io.BytesIO(template_bytes))
    except Exception as e:
        raise PopulateError(f"could not load template workbook: {e}") from e

    address_by_path = _flatten_cellmap(cellmap)

    for path, spec in assumptions.items():
        addr = address_by_path.get(path)
        if addr is None:
            # Unknown canonical path -- silently skip. The cellmap may simply
            # not carry this field for this template.
            continue
        if not isinstance(spec, dict) or "base" not in spec:
            raise PopulateError(
                f"assumption for {path!r} must be a dict with at least a 'base' key"
            )
        value = spec["base"]
        cell = _resolve_address(wb, addr)
        try:
            cell.value = value
        except (ValueError, TypeError) as e:
            raise PopulateError(
                f"could not write value for {path!r} at {addr}: {e}"
            ) from e

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
